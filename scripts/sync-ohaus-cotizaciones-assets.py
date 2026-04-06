import argparse
import json
import mimetypes
import os
import re
from pathlib import Path

import openpyxl
import requests


SYSTEM_USER_ID = "841263c6-196d-49cd-b5ba-aae0b097014f"
DEFAULT_PROVIDER = "ohaus_colombia"
DEFAULT_FOLDER = "app/api/agents/channels/evolution/webhook/Ohaus/Cotizaciones"
DEFAULT_BUCKET = "ohaus-cotizaciones"


MODEL_RE = re.compile(r"\b([A-Z]{1,5}\d[\dA-Z/-]{0,14})\b")


def load_env_file(path: Path):
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def load_env():
    root = Path.cwd()
    load_env_file(root / ".env.local")
    load_env_file(root / ".env")


def norm_model(v: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(v or "").upper())


def model_from_filename(file_name: str) -> str:
    stem = Path(file_name).stem
    clean = re.sub(r"^\d+\.", "", stem).strip()
    clean = re.sub(r"\b(ficha|tecnica|tecnico|datasheet|data\s*sheet|cotizacion|cotizacion)\b", " ", clean, flags=re.I)
    clean = re.sub(r"\s+", " ", clean).strip(" -_")
    found = MODEL_RE.findall(clean.upper())
    if found:
        ranked = sorted(found, key=lambda x: (len(norm_model(x)), x.count("-"), x), reverse=True)
        return ranked[0]
    parts = [p for p in re.split(r"\s+", clean) if p]
    return parts[-1].upper() if parts else ""


def prefix_id_from_filename(file_name: str) -> str:
    m = re.match(r"^(\d+)\.", str(file_name or "").strip())
    return m.group(1) if m else ""


def model_from_xlsx_cell(xlsx_path: Path) -> str:
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        candidates = [
            str(ws.cell(5, 2).value or "").strip(),
            str(ws.cell(6, 2).value or "").strip(),
            str(ws.cell(5, 4).value or "").strip(),
        ]
        wb.close()
    except Exception:
        return ""
    for c in candidates:
        m = MODEL_RE.findall(c.upper())
        if m:
            return m[0]
    return ""


def pick_embedded_product_image(xlsx_path: Path):
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        images = list(getattr(ws, "_images", []) or [])
        if not images:
            wb.close()
            return None

        best = None
        best_score = -10**9
        for img in images:
            w = int(getattr(img, "width", 0) or 0)
            h = int(getattr(img, "height", 0) or 0)
            if w <= 0 or h <= 0:
                continue

            anc = getattr(img, "anchor", None)
            fr = getattr(anc, "_from", None)
            row = int(getattr(fr, "row", 0) or 0)
            col = int(getattr(fr, "col", 0) or 0)

            area = w * h
            ratio = w / max(1, h)
            score = 0
            if 3 <= row <= 16:
                score += 60
            if col <= 4:
                score += 30
            if 0.55 <= ratio <= 1.9:
                score += 35
            if w >= 220 and h >= 220:
                score += 20
            if w > 1000 and h < 700:
                score -= 120
            if row <= 1:
                score -= 50
            score += min(40, area // 6000)

            if score > best_score:
                best_score = score
                best = img

        if not best:
            wb.close()
            return None

        raw = best._data()
        ext = str(getattr(best, "path", "") or "").lower()
        if ext.endswith(".jpg") or ext.endswith(".jpeg"):
            suffix = ".jpg"
        elif ext.endswith(".webp"):
            suffix = ".webp"
        else:
            suffix = ".png"
        wb.close()
        return {"bytes": raw, "ext": suffix, "score": best_score}
    except Exception:
        return None


def choose_best_pdf(files):
    if not files:
        return None
    ranked = []
    for p in files:
        name = p.name.lower()
        score = 0
        if "ficha" in name:
            score += 20
        if "datasheet" in name or "data sheet" in name:
            score += 12
        if "manual" in name or "brochure" in name:
            score -= 10
        score -= len(name) // 40
        ranked.append((score, p))
    ranked.sort(key=lambda x: x[0], reverse=True)
    return ranked[0][1]


def build_assets_index(folder: Path):
    by_key = {}
    key_by_prefix = {}
    for x in sorted(folder.glob("*.xlsx")):
        model = model_from_xlsx_cell(x) or model_from_filename(x.name)
        key = norm_model(model)
        if not key:
            continue
        if key not in by_key:
            by_key[key] = {"model": model, "xlsx": None, "pdfs": []}
        by_key[key]["xlsx"] = x
        pref = prefix_id_from_filename(x.name)
        if pref:
            key_by_prefix[pref] = key

    for p in sorted(folder.glob("*.pdf")):
        pref = prefix_id_from_filename(p.name)
        key = key_by_prefix.get(pref, "") if pref else ""
        model = ""
        if not key:
            model = model_from_filename(p.name)
            key = norm_model(model)
        if not key:
            continue
        if key not in by_key:
            by_key[key] = {"model": model, "xlsx": None, "pdfs": []}
        by_key[key]["pdfs"].append(p)

    out = {}
    for key, row in by_key.items():
        pdf = choose_best_pdf(row.get("pdfs") or [])
        out[key] = {
            "model": row.get("model") or key,
            "xlsx": row.get("xlsx"),
            "pdf": pdf,
        }
    return out


def request_ok(resp: requests.Response):
    if 200 <= resp.status_code < 300:
        return
    raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:800]}")


def ensure_bucket(base_url: str, service_key: str, bucket: str):
    url = f"{base_url}/storage/v1/bucket"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
    }
    payload = {"id": bucket, "name": bucket, "public": True}
    resp = requests.post(url, headers=headers, data=json.dumps(payload).encode("utf-8"), timeout=60)
    if resp.status_code in (200, 201):
        return
    if resp.status_code in (400, 409):
        return
    request_ok(resp)


def upload_object(base_url: str, service_key: str, bucket: str, object_path: str, blob: bytes, mime_type: str):
    object_path = object_path.replace("\\", "/").lstrip("/")
    url = f"{base_url}/storage/v1/object/{bucket}/{object_path}"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": mime_type,
        "x-upsert": "true",
    }
    resp = requests.post(url, headers=headers, data=blob, timeout=180)
    if resp.status_code not in (200, 201):
        request_ok(resp)
    return f"{base_url}/storage/v1/object/public/{bucket}/{object_path}"


def get_catalog_rows(base_url: str, service_key: str, owner_id: str, provider: str):
    url = f"{base_url}/rest/v1/agent_product_catalog"
    headers = {"apikey": service_key, "Authorization": f"Bearer {service_key}"}
    params = {
        "select": "id,name,source_payload,image_url,datasheet_url",
        "created_by": f"eq.{owner_id}",
        "provider": f"eq.{provider}",
        "limit": "10000",
    }
    resp = requests.get(url, headers=headers, params=params, timeout=90)
    request_ok(resp)
    return resp.json() if resp.text else []


def patch_catalog_row(base_url: str, service_key: str, row_id: str, patch_obj: dict):
    url = f"{base_url}/rest/v1/agent_product_catalog"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    params = {"id": f"eq.{row_id}"}
    resp = requests.patch(
        url,
        headers=headers,
        params=params,
        data=json.dumps(patch_obj, ensure_ascii=False).encode("utf-8"),
        timeout=90,
    )
    request_ok(resp)


def main():
    load_env()
    parser = argparse.ArgumentParser(description="Upload Ohaus Cotizaciones assets and patch catalog image_url/datasheet_url")
    parser.add_argument("--folder", default=os.getenv("OHAUS_COTIZACIONES_FOLDER", DEFAULT_FOLDER))
    parser.add_argument("--bucket", default=os.getenv("OHAUS_COTIZACIONES_BUCKET", DEFAULT_BUCKET))
    parser.add_argument("--owner-id", default=os.getenv("QUOTE_TEMPLATE_OWNER_ID", SYSTEM_USER_ID))
    parser.add_argument("--provider", default=os.getenv("QUOTE_TEMPLATE_PROVIDER", DEFAULT_PROVIDER))
    parser.add_argument("--apply", action="store_true", help="Apply changes (default dry-run)")
    args = parser.parse_args()

    base_url = (os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL") or "").strip().rstrip("/")
    service_key = (os.getenv("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
    if not base_url or not service_key:
        raise RuntimeError("Missing SUPABASE_URL/NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")

    folder = Path(args.folder)
    if not folder.is_absolute():
        folder = Path.cwd() / folder
    if not folder.exists():
        raise RuntimeError(f"Folder not found: {folder}")

    index = build_assets_index(folder)
    catalog = get_catalog_rows(base_url, service_key, args.owner_id, args.provider)

    print(f"Folder: {folder}")
    print(f"Models in folder index: {len(index)}")
    print(f"Catalog rows: {len(catalog)}")

    if args.apply:
        ensure_bucket(base_url, service_key, args.bucket)

    updates = []
    misses = 0
    missed_models = []
    for row in catalog:
        model = str(row.get("name") or "").strip()
        key = norm_model(model)
        assets = index.get(key)
        if not assets:
            misses += 1
            missed_models.append(model)
            continue

        pdf_path = assets.get("pdf")
        xlsx_path = assets.get("xlsx")
        next_datasheet_url = ""
        next_image_url = ""

        if args.apply and pdf_path and pdf_path.exists():
            pdf_bytes = pdf_path.read_bytes()
            next_datasheet_url = upload_object(
                base_url,
                service_key,
                args.bucket,
                f"datasheets/{key}.pdf",
                pdf_bytes,
                "application/pdf",
            )
        elif pdf_path:
            next_datasheet_url = f"(dry-run) datasheets/{key}.pdf"

        img = pick_embedded_product_image(xlsx_path) if xlsx_path and xlsx_path.exists() else None
        if args.apply and img:
            ext = str(img.get("ext") or ".png")
            mime = mimetypes.types_map.get(ext.lower(), "image/png")
            next_image_url = upload_object(
                base_url,
                service_key,
                args.bucket,
                f"images/{key}{ext}",
                img.get("bytes") or b"",
                mime,
            )
        elif img:
            next_image_url = f"(dry-run) images/{key}{str(img.get('ext') or '.png')}"

        patch_obj = {}
        if next_datasheet_url:
            patch_obj["datasheet_url"] = next_datasheet_url
        if next_image_url:
            patch_obj["image_url"] = next_image_url

        source_payload = row.get("source_payload") if isinstance(row.get("source_payload"), dict) else {}
        if patch_obj:
            patch_obj["source_payload"] = {
                **source_payload,
                "import_source": "Ohaus/Cotizaciones assets",
                "quote_model": assets.get("model") or model,
                "quote_xlsx_file": xlsx_path.name if xlsx_path else source_payload.get("quote_xlsx_file"),
                "quote_pdf_file": pdf_path.name if pdf_path else source_payload.get("quote_pdf_file"),
            }
            updates.append((str(row.get("id")), model, patch_obj, bool(pdf_path), bool(img)))

    print(f"Catalog rows matched: {len(catalog) - misses}")
    print(f"Catalog rows without folder match: {misses}")
    if missed_models:
        print("Missing models:", ", ".join(sorted(missed_models)[:12]))
    print(f"Rows prepared to update: {len(updates)}")
    print(f"Rows with PDF: {sum(1 for _, _, _, has_pdf, _ in updates if has_pdf)}")
    print(f"Rows with image from XLSX: {sum(1 for _, _, _, _, has_img in updates if has_img)}")

    if updates:
        sample = updates[0]
        print("Sample model:", sample[1])

    if not args.apply:
        print("Dry run mode. Use --apply to upload and patch DB.")
        return

    for row_id, _model, patch_obj, _has_pdf, _has_img in updates:
        patch_catalog_row(base_url, service_key, row_id, patch_obj)

    print(f"Updated rows: {len(updates)}")


if __name__ == "__main__":
    main()
