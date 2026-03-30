import json
import os
import re
from pathlib import Path

import openpyxl
import requests


SYSTEM_USER_ID = "841263c6-196d-49cd-b5ba-aae0b097014f"
DEFAULT_PROVIDER = "ohaus_colombia"
DEFAULT_XLSX = "app/api/agents/channels/evolution/webhook-v2/plantilla_descripcion_cotizaciones_ohausfinal.xlsx"


def load_env_file(path: Path):
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def load_env():
    root = Path.cwd()
    load_env_file(root / ".env.local")
    load_env_file(root / ".env")


def norm(s: str) -> str:
    s = str(s or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "", s)


def clean(v):
    s = str(v or "").strip()
    if s in {"-", "--", "n/a", "na", "none", "null"}:
        return ""
    return s


def num(v):
    s = clean(v)
    if not s:
        return None
    s = re.sub(r"[^0-9,.-]", "", s)
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        n = float(s)
    except Exception:
        return None
    return round(n, 2) if n > 0 else None


def parse_template_rows(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [clean(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    rows = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        model = clean(data.get("modelo", ""))
        if not model:
            continue
        key = norm(model)
        if not key:
            continue

        desc_short = clean(data.get("descripcion_comercial_corta", ""))
        desc_long = clean(data.get("descripcion_comercial_larga", ""))
        family = clean(data.get("familia", ""))
        brand = clean(data.get("marca", ""))
        image_url = clean(data.get("imagen_url", ""))
        ficha_pdf_url = clean(data.get("ficha_pdf_url", ""))
        price_bogota = num(data.get("precio_bogota_cop"))
        price_antioquia = num(data.get("precio_antioquia_cop"))
        price_dist = num(data.get("precio_distribuidor_cop"))

        rows[key] = {
            "modelo": model,
            "summary": desc_short,
            "description": desc_long or desc_short,
            "family": family,
            "brand": brand,
            "image_url": image_url,
            "datasheet_url": ficha_pdf_url,
            "prices_cop": {
                "bogota": price_bogota,
                "antioquia": price_antioquia,
                "distribuidor": price_dist,
            },
            "_row": i,
        }
    return rows


def request_ok(resp: requests.Response):
    if 200 <= resp.status_code < 300:
        return
    raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:800]}")


def get_catalog_rows(base_url: str, service_key: str, owner_id: str, provider: str):
    url = f"{base_url}/rest/v1/agent_product_catalog"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
    }
    params = {
        "select": "id,name,source_payload,image_url,datasheet_url,summary,description",
        "created_by": f"eq.{owner_id}",
        "provider": f"eq.{provider}",
        "limit": "5000",
    }
    resp = requests.get(url, headers=headers, params=params, timeout=120)
    request_ok(resp)
    return resp.json() if resp.text else []


def patch_catalog_row(base_url: str, service_key: str, row_id: str, patch_obj: dict):
    url = f"{base_url}/rest/v1/agent_product_catalog"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    params = {"id": f"eq.{row_id}"}
    resp = requests.patch(url, headers=headers, params=params, data=json.dumps(patch_obj, ensure_ascii=False).encode("utf-8"), timeout=120)
    request_ok(resp)


def main():
    load_env()
    base_url = (os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL") or "").strip().rstrip("/")
    service_key = (os.getenv("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
    owner_id = (os.getenv("QUOTE_TEMPLATE_OWNER_ID") or SYSTEM_USER_ID).strip()
    provider = (os.getenv("QUOTE_TEMPLATE_PROVIDER") or DEFAULT_PROVIDER).strip()
    dry_run = (os.getenv("QUOTE_TEMPLATE_DRY_RUN") or "false").strip().lower() in {"1", "true", "yes"}

    xlsx_path = Path(os.getenv("QUOTE_TEMPLATE_XLSX") or DEFAULT_XLSX)
    if not xlsx_path.is_absolute():
        xlsx_path = Path.cwd() / xlsx_path
    if not xlsx_path.exists():
        raise RuntimeError(f"Template XLSX not found: {xlsx_path}")
    if not base_url or not service_key:
        raise RuntimeError("Missing SUPABASE_URL/NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")

    template = parse_template_rows(xlsx_path)
    catalog = get_catalog_rows(base_url, service_key, owner_id, provider)

    updates = []
    misses = 0
    for row in catalog:
        model = str(row.get("name") or "").strip()
        key = norm(model)
        tpl = template.get(key)
        if not tpl:
            misses += 1
            continue
        source_payload = row.get("source_payload") if isinstance(row.get("source_payload"), dict) else {}
        prices = source_payload.get("prices_cop") if isinstance(source_payload.get("prices_cop"), dict) else {}
        tpl_prices = tpl.get("prices_cop") or {}
        merged_prices = {
            "bogota": tpl_prices.get("bogota") if tpl_prices.get("bogota") is not None else prices.get("bogota"),
            "antioquia": tpl_prices.get("antioquia") if tpl_prices.get("antioquia") is not None else prices.get("antioquia"),
            "distribuidor": tpl_prices.get("distribuidor") if tpl_prices.get("distribuidor") is not None else prices.get("distribuidor"),
        }
        merged_source = {
            **source_payload,
            "family": tpl.get("family") or source_payload.get("family"),
            "quote_description": tpl.get("description") or source_payload.get("quote_description"),
            "descripcion_comercial_corta": tpl.get("summary") or source_payload.get("descripcion_comercial_corta"),
            "descripcion_comercial_larga": tpl.get("description") or source_payload.get("descripcion_comercial_larga"),
            "prices_cop": merged_prices,
        }
        patch_obj = {
            "summary": tpl.get("summary") or row.get("summary"),
            "description": tpl.get("description") or row.get("description"),
            "source_payload": merged_source,
        }
        if tpl.get("image_url"):
            patch_obj["image_url"] = tpl.get("image_url")
        if tpl.get("datasheet_url"):
            patch_obj["datasheet_url"] = tpl.get("datasheet_url")
        updates.append((str(row.get("id")), patch_obj, model))

    print(f"Template models: {len(template)}")
    print(f"Catalog rows: {len(catalog)}")
    print(f"Rows to update: {len(updates)}")
    print(f"Catalog rows without template match: {misses}")
    if updates:
        print("Sample update model:", updates[0][2])

    if dry_run:
        print("Dry run enabled, no DB write executed.")
        return

    for row_id, patch_obj, _model in updates:
        patch_catalog_row(base_url, service_key, row_id, patch_obj)

    print(f"Updated rows: {len(updates)}")


if __name__ == "__main__":
    main()
