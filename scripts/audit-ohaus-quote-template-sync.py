import argparse
import os
import re
from pathlib import Path

import openpyxl
import requests


SYSTEM_USER_ID = "841263c6-196d-49cd-b5ba-aae0b097014f"
DEFAULT_PROVIDER = "ohaus_colombia"
DEFAULT_XLSX = "app/api/agents/channels/evolution/webhook-v2/plantilla_descripcion_cotizaciones_ohausfinal.xlsx"


def load_env_file(path: Path):
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def load_env():
    root = Path.cwd()
    load_env_file(root / ".env.local")
    load_env_file(root / ".env")


def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s or "").strip().lower())


def clean(v):
    s = str(v or "").strip()
    if s.lower() in {"-", "--", "n/a", "na", "none", "null"}:
        return ""
    return s


def num(v):
    s = clean(v)
    if not s:
        return None
    s = re.sub(r"[^0-9,.-]", "", s)
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        out = float(s)
    except Exception:
        return None
    return round(out, 2)


def parse_template_rows(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [clean(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        model = clean(data.get("modelo", ""))
        if not model:
            continue
        key = norm(model)
        if not key:
            continue
        rows[key] = {
            "model": model,
            "summary": clean(data.get("descripcion_comercial_corta", "")),
            "description": clean(data.get("descripcion_comercial_larga", "")) or clean(data.get("descripcion_comercial_corta", "")),
            "family": clean(data.get("familia", "")),
            "image_url": clean(data.get("imagen_url", "")),
            "datasheet_url": clean(data.get("ficha_pdf_url", "")),
            "prices_cop": {
                "bogota": num(data.get("precio_bogota_cop")),
                "antioquia": num(data.get("precio_antioquia_cop")),
                "distribuidor": num(data.get("precio_distribuidor_cop")),
            },
        }
    wb.close()
    return rows


def request_ok(resp: requests.Response):
    if 200 <= resp.status_code < 300:
        return
    raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:800]}")


def get_catalog_rows(base_url: str, service_key: str, owner_id: str, provider: str):
    url = f"{base_url}/rest/v1/agent_product_catalog"
    headers = {"apikey": service_key, "Authorization": f"Bearer {service_key}"}
    params = {
        "select": "id,name,summary,description,image_url,datasheet_url,source_payload",
        "created_by": f"eq.{owner_id}",
        "provider": f"eq.{provider}",
        "limit": "5000",
    }
    resp = requests.get(url, headers=headers, params=params, timeout=120)
    request_ok(resp)
    return resp.json() if resp.text else []


def main():
    load_env()
    parser = argparse.ArgumentParser(description="Audit XLS template sync against agent_product_catalog")
    parser.add_argument("--xlsx", default=os.getenv("QUOTE_TEMPLATE_XLSX") or DEFAULT_XLSX)
    parser.add_argument("--owner-id", default=os.getenv("QUOTE_TEMPLATE_OWNER_ID") or SYSTEM_USER_ID)
    parser.add_argument("--provider", default=os.getenv("QUOTE_TEMPLATE_PROVIDER") or DEFAULT_PROVIDER)
    parser.add_argument("--max-errors", type=int, default=40)
    args = parser.parse_args()

    base_url = (os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL") or "").strip().rstrip("/")
    service_key = (os.getenv("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
    if not base_url or not service_key:
        raise RuntimeError("Missing SUPABASE_URL/NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = Path.cwd() / xlsx_path
    if not xlsx_path.exists():
        raise RuntimeError(f"Template XLSX not found: {xlsx_path}")

    template = parse_template_rows(xlsx_path)
    catalog_rows = get_catalog_rows(base_url, service_key, str(args.owner_id).strip(), str(args.provider).strip())
    catalog_by_key = {norm(str(r.get("name") or "")): r for r in catalog_rows if norm(str(r.get("name") or ""))}

    missing_models = []
    mismatches = []

    for key, tpl in template.items():
        row = catalog_by_key.get(key)
        if not row:
            missing_models.append(tpl.get("model") or key)
            continue

        payload = row.get("source_payload") if isinstance(row.get("source_payload"), dict) else {}
        checks = {
            "summary": (clean(tpl.get("summary")), clean(row.get("summary"))),
            "description": (clean(tpl.get("description")), clean(row.get("description"))),
            "family": (clean(tpl.get("family")), clean(payload.get("family"))),
            "image_url": (clean(tpl.get("image_url")), clean(row.get("image_url"))),
            "datasheet_url": (clean(tpl.get("datasheet_url")), clean(row.get("datasheet_url"))),
            "price_bogota": (tpl.get("prices_cop", {}).get("bogota"), num((payload.get("prices_cop") or {}).get("bogota"))),
            "price_antioquia": (tpl.get("prices_cop", {}).get("antioquia"), num((payload.get("prices_cop") or {}).get("antioquia"))),
            "price_distribuidor": (tpl.get("prices_cop", {}).get("distribuidor"), num((payload.get("prices_cop") or {}).get("distribuidor"))),
        }

        for field, (expected, got) in checks.items():
            if expected in (None, ""):
                continue
            if isinstance(expected, float) or isinstance(got, float):
                if expected is None and got is None:
                    continue
                if float(expected or 0) != float(got or 0):
                    mismatches.append(f"{tpl.get('model')}: {field} expected={expected} got={got}")
            else:
                if norm(str(expected)) != norm(str(got)):
                    mismatches.append(f"{tpl.get('model')}: {field} expected='{expected}' got='{got}'")

    print(f"Template models: {len(template)}")
    print(f"Catalog rows: {len(catalog_rows)}")
    print(f"Missing models in DB: {len(missing_models)}")
    print(f"Field mismatches: {len(mismatches)}")

    shown = 0
    for m in missing_models:
        if shown >= args.max_errors:
            break
        print(f"MISSING: {m}")
        shown += 1

    for mm in mismatches:
        if shown >= args.max_errors:
            break
        print(f"MISMATCH: {mm}")
        shown += 1

    if missing_models or mismatches:
        raise SystemExit(1)

    print("Audit OK: XLS and DB are aligned.")


if __name__ == "__main__":
    main()
