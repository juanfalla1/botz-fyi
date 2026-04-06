import argparse
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import requests


SYSTEM_USER_ID = "841263c6-196d-49cd-b5ba-aae0b097014f"
DEFAULT_PROVIDER = "ohaus_colombia"
DEFAULT_TRM = 4200.0
DEFAULT_PRICE_FILE = "app/api/agents/channels/evolution/webhook-v2/Lista de precios ohaus IA.xlsx"


def load_env_file(path: Path):
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def load_env():
    root = Path.cwd()
    load_env_file(root / ".env.local")
    load_env_file(root / ".env")


def norm_model(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def parse_number(value):
    s = str(value or "").strip()
    if not s:
        return None
    s = re.sub(r"[^0-9,.-]", "", s)
    if not s:
        return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        n = float(s)
    except Exception:
        return None
    return round(n, 2) if n > 0 else None


def parse_price_list_xlsx(file_path: Path):
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    out = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        family = str(row[0] or "").strip()
        model = str(row[1] or "").strip()
        capacity = str(row[2] or "").strip()
        resolution = str(row[3] or "").strip()
        antioquia = parse_number(row[4])
        bogota = parse_number(row[5])
        dist = parse_number(row[6])

        if not model:
            continue

        key = norm_model(model)
        if not key:
            continue

        if not (antioquia or bogota or dist):
            continue

        out[key] = {
            "model": model,
            "family": family,
            "capacity": capacity,
            "resolution": resolution,
            "prices_cop": {
                "antioquia": antioquia,
                "bogota": bogota,
                "distribuidor": dist,
            },
            "price_cop_selected": bogota or antioquia or dist,
        }

    wb.close()
    return out


def request_ok(resp: requests.Response):
    if 200 <= resp.status_code < 300:
        return
    raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:700]}")


def get_catalog_rows(base_url: str, service_key: str, owner_id: str, provider: str, timeout_sec: float):
    url = f"{base_url}/rest/v1/agent_product_catalog"
    headers = {"apikey": service_key, "Authorization": f"Bearer {service_key}"}
    params = {
        "select": "id,name,source_payload,base_price_usd,price_currency",
        "created_by": f"eq.{owner_id}",
        "provider": f"eq.{provider}",
        "limit": "10000",
    }
    resp = requests.get(url, headers=headers, params=params, timeout=(10, timeout_sec))
    request_ok(resp)
    return resp.json() if resp.text else []


def patch_catalog_row(base_url: str, service_key: str, row_id: str, patch_obj: dict, timeout_sec: float):
    url = f"{base_url}/rest/v1/agent_product_catalog"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    params = {"id": f"eq.{row_id}"}
    resp = requests.patch(
        url,
        headers=headers,
        params=params,
        data=json.dumps(patch_obj, ensure_ascii=False).encode("utf-8"),
        timeout=(10, timeout_sec),
    )
    request_ok(resp)


def main():
    load_env()

    parser = argparse.ArgumentParser(description="Sync Ohaus prices from Excel list into catalog")
    parser.add_argument("--file", default=os.getenv("OHAUS_PRICE_FILE", DEFAULT_PRICE_FILE))
    parser.add_argument("--owner-id", default=os.getenv("QUOTE_TEMPLATE_OWNER_ID", SYSTEM_USER_ID))
    parser.add_argument("--provider", default=os.getenv("QUOTE_TEMPLATE_PROVIDER", DEFAULT_PROVIDER))
    parser.add_argument("--trm", type=float, default=float(os.getenv("OHAUS_SYNC_TRM") or DEFAULT_TRM))
    parser.add_argument("--http-timeout", type=float, default=float(os.getenv("OHAUS_SYNC_HTTP_TIMEOUT") or 35))
    parser.add_argument("--apply", action="store_true", help="Apply updates to DB. Default is dry-run")
    args = parser.parse_args()

    base_url = (os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL") or "").strip().rstrip("/")
    service_key = (os.getenv("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
    if not base_url or not service_key:
        raise RuntimeError("Missing SUPABASE_URL/NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")

    file_path = Path(args.file)
    if not file_path.is_absolute():
        file_path = Path.cwd() / file_path
    if not file_path.exists():
        raise RuntimeError(f"Price file not found: {file_path}")

    prices = parse_price_list_xlsx(file_path)
    if not prices:
        raise RuntimeError("No prices parsed from Excel")

    catalog = get_catalog_rows(base_url, service_key, args.owner_id, args.provider, args.http_timeout)
    now_iso = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

    updates = []
    misses = 0

    for row in catalog:
        model = str(row.get("name") or "").strip()
        key = norm_model(model)
        incoming = prices.get(key)
        if not incoming:
            misses += 1
            continue

        price_cop = float(incoming.get("price_cop_selected") or 0)
        if price_cop <= 0:
            continue

        source_payload = row.get("source_payload") if isinstance(row.get("source_payload"), dict) else {}
        old_prices = source_payload.get("prices_cop") if isinstance(source_payload.get("prices_cop"), dict) else {}
        new_prices = {**old_prices, **incoming.get("prices_cop", {})}

        merged_source = {
            **source_payload,
            "import_source": "Lista de precios ohaus IA.xlsx",
            "prices_cop": new_prices,
            "family": incoming.get("family") or source_payload.get("family"),
            "capacity": incoming.get("capacity") or source_payload.get("capacity"),
            "resolution": incoming.get("resolution") or source_payload.get("resolution"),
            "price_list_synced_at": now_iso,
        }

        patch_obj = {
            "source_payload": merged_source,
            "base_price_usd": round(price_cop / args.trm, 6),
            "price_currency": "USD",
            "last_price_update": now_iso,
        }

        updates.append((str(row.get("id")), model, patch_obj, incoming))

    print(f"Price file: {file_path}")
    print(f"Price models parsed: {len(prices)}")
    print(f"Catalog rows: {len(catalog)}")
    print(f"Catalog rows matched for price update: {len(updates)}")
    print(f"Catalog rows without price match: {misses}")

    if updates:
        sample = updates[0]
        print(
            "Sample:",
            json.dumps(
                {
                    "catalog_model": sample[1],
                    "price_bogota": sample[3].get("prices_cop", {}).get("bogota"),
                    "base_price_usd": sample[2].get("base_price_usd"),
                },
                ensure_ascii=False,
            ),
        )

    if not args.apply:
        print("Dry run mode. Use --apply to write changes.")
        return

    for row_id, _model, patch_obj, _incoming in updates:
        patch_catalog_row(base_url, service_key, row_id, patch_obj, args.http_timeout)

    print(f"Updated rows: {len(updates)}")


if __name__ == "__main__":
    main()
