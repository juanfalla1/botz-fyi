import json
import os
import re
from pathlib import Path

import openpyxl
import requests


SYSTEM_TENANT_ID = "0811c118-5a2f-40cb-907e-8979e0984096"
SYSTEM_USER_ID = "841263c6-196d-49cd-b5ba-aae0b097014f"
DEFAULT_PROVIDER = "ohaus_colombia"
DEFAULT_XLSX = "app/api/agents/channels/evolution/webhook/Lista de precios ohaus IA y productos .xlsx"


def load_env_file(path: Path):
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def load_env():
    root = Path.cwd()
    load_env_file(root / ".env.local")
    load_env_file(root / ".env")


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip().lower())


def slugify(text: str) -> str:
    t = normalize(text)
    t = re.sub(r"[^a-z0-9]+", "-", t)
    return t.strip("-")[:140]


def category_from_family(family: str) -> str:
    t = normalize(family)
    if "humedad" in t:
        return "analizador_humedad"
    if any(x in t for x in ["bascula", "ranger", "defender", "valor", "plataforma", "indicador"]):
        return "basculas"
    if any(x in t for x in ["electrodo", "ph", "orp", "conductividad", "electroquim"]):
        return "electroquimica"
    if "impresora" in t:
        return "impresoras"
    if any(x in t for x in ["centrifuga", "agitador", "mezclador", "homogeneizador", "laboratorio", "plancha"]):
        return "equipos_laboratorio"
    return "balanzas"


def num(v):
    if v is None:
        return None
    s = str(v).strip().replace(" ", "")
    if not s:
        return None
    s = re.sub(r"[^0-9,.-]", "", s)
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        n = float(s)
    except Exception:
        return None
    return round(n, 2) if n > 0 else None


def parse_xlsx(xlsx_path: Path, tenant_id: str, created_by: str, provider: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    seen_urls = set()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        family = str(row[0] or "").strip()
        model = str(row[1] or "").strip()
        capacity = str(row[2] or "").strip()
        resolution = str(row[3] or "").strip()
        price_antioquia = num(row[4])
        price_bogota = num(row[5])
        price_dist = num(row[6])

        if not model:
            continue

        if normalize(model) in ["modelo", "none"]:
            continue

        category = category_from_family(family)
        product_url = f"https://catalogo.ohaus.local/modelo/{model}"
        if product_url in seen_urls:
            product_url = f"{product_url}-{i}"
        seen_urls.add(product_url)

        price_ref = price_bogota or price_antioquia or price_dist or 0
        approx_usd = round(price_ref / 4200, 6) if price_ref > 0 else 0

        specs_text = f"Familia: {family}; Capacidad: {capacity}; Resolucion: {resolution}"
        rows.append(
            {
                "tenant_id": tenant_id,
                "created_by": created_by,
                "provider": provider,
                "brand": "OHAUS",
                "category": category,
                "name": model,
                "slug": slugify(model),
                "product_url": product_url,
                "image_url": None,
                "summary": f"{family} | Capacidad {capacity} | Resolucion {resolution}"[:500],
                "description": specs_text,
                "standards": [],
                "methods": [],
                "specs_text": specs_text,
                "specs_json": {
                    "familia": family,
                    "modelo": model,
                    "capacidad": capacity,
                    "resolucion": resolution,
                },
                "source_payload": {
                    "import_source": "Lista de precios ohaus IA y productos .xlsx",
                    "family": family,
                    "model": model,
                    "capacity": capacity,
                    "resolution": resolution,
                    "prices_cop": {
                        "antioquia": price_antioquia,
                        "bogota": price_bogota,
                        "distribuidor": price_dist,
                    },
                },
                "is_active": True,
                "base_price_usd": approx_usd,
                "price_currency": "USD",
                "last_price_update": None,
                "datasheet_url": None,
            }
        )

    return rows


def supabase_headers(service_key: str):
    return {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }


def request_ok(resp: requests.Response):
    if 200 <= resp.status_code < 300:
        return
    raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:400]}")


def chunked(arr, size):
    for i in range(0, len(arr), size):
        yield arr[i : i + size]


def main():
    load_env()
    supabase_url = (os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
    service_key = (os.getenv("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
    if not supabase_url or not service_key:
        raise RuntimeError("Missing SUPABASE_URL/NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")

    tenant_id = (os.getenv("CATALOG_IMPORT_TENANT_ID") or SYSTEM_TENANT_ID).strip()
    created_by = (os.getenv("CATALOG_IMPORT_USER_ID") or SYSTEM_USER_ID).strip()
    provider = (os.getenv("CATALOG_IMPORT_PROVIDER") or DEFAULT_PROVIDER).strip()
    xlsx_path = Path(os.getenv("CATALOG_IMPORT_XLSX") or DEFAULT_XLSX)
    if not xlsx_path.is_absolute():
        xlsx_path = Path.cwd() / xlsx_path
    if not xlsx_path.exists():
        raise RuntimeError(f"XLSX file not found: {xlsx_path}")

    rows = parse_xlsx(xlsx_path, tenant_id, created_by, provider)
    if not rows:
        raise RuntimeError("No products parsed from XLSX")

    print(f"Parsed rows: {len(rows)}")
    print("Sample:", json.dumps(rows[:3], ensure_ascii=False)[:700])

    base = supabase_url.rstrip("/") + "/rest/v1"
    headers = supabase_headers(service_key)

    # Read old ids
    sel_params = {
        "select": "id",
        "tenant_id": f"eq.{tenant_id}",
        "created_by": f"eq.{created_by}",
        "provider": f"eq.{provider}",
        "limit": "20000",
    }
    r = requests.get(f"{base}/agent_product_catalog", headers=headers, params=sel_params, timeout=60)
    request_ok(r)
    old_ids = [x.get("id") for x in (r.json() or []) if x.get("id")]

    # Delete variants first
    if old_ids:
        for batch in chunked(old_ids, 300):
            p = {"catalog_id": f"in.({','.join(batch)})"}
            rv = requests.delete(f"{base}/agent_product_variants", headers=headers, params=p, timeout=60)
            if rv.status_code >= 400 and "relation" not in (rv.text or "").lower():
                raise RuntimeError(f"Failed deleting variants: {rv.status_code} {rv.text[:300]}")

    # Delete old catalog
    del_params = {
        "tenant_id": f"eq.{tenant_id}",
        "created_by": f"eq.{created_by}",
        "provider": f"eq.{provider}",
    }
    rd = requests.delete(f"{base}/agent_product_catalog", headers=headers, params=del_params, timeout=120)
    request_ok(rd)

    # Insert new rows
    inserted = 0
    for batch in chunked(rows, 250):
        ri = requests.post(f"{base}/agent_product_catalog", headers=headers, json=batch, timeout=120)
        request_ok(ri)
        inserted += len(batch)
        print(f"Inserted {inserted}/{len(rows)}")

    print("Import completed")


if __name__ == "__main__":
    main()
