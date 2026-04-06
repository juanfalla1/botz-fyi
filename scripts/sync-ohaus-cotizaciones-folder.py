import argparse
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import requests


SYSTEM_USER_ID = "841263c6-196d-49cd-b5ba-aae0b097014f"
DEFAULT_PROVIDER = "ohaus_colombia"
DEFAULT_FOLDER = "app/api/agents/channels/evolution/webhook/Ohaus/Cotizaciones"
DEFAULT_TRM = 4200.0


def load_env_file(path: Path):
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def load_env():
    root = Path.cwd()
    load_env_file(root / ".env.local")
    load_env_file(root / ".env")


def norm_text(v: str) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip())


def norm_model(v: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(v or "").upper())


def parse_money(value) -> float | None:
    s = str(value or "").strip()
    if not s:
        return None
    s = re.sub(r"[^0-9,.-]", "", s)
    if not s:
        return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        n = float(s)
    except Exception:
        return None
    return round(n, 2) if n > 0 else None


MODEL_RE = re.compile(r"\b([A-Z]{1,5}\d[\dA-Z/-]{0,12})\b")


def model_from_filename(file_name: str) -> str:
    stem = Path(file_name).stem
    clean = re.sub(r"^\d+\.", "", stem).strip()
    found = MODEL_RE.findall(clean.upper())
    if found:
        return found[-1]
    tokens = [t for t in re.split(r"\s+", clean) if t]
    if not tokens:
        return ""
    return tokens[-1].upper()


def iter_sheet_cells(ws, max_rows=180, max_cols=14):
    for r, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=True),
        start=1,
    ):
        row_vals = []
        for c, v in enumerate(row, start=1):
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            row_vals.append((c, s))
        if row_vals:
            yield r, row_vals


def parse_xlsx_product(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    model_candidates = []
    description_candidates = []
    price_candidates = []

    for _r, row_vals in iter_sheet_cells(ws):
        row_text = " | ".join(v for _, v in row_vals)

        for _c, value in row_vals:
            up = value.upper()
            for m in MODEL_RE.findall(up):
                if m.startswith("SAP"):
                    continue
                model_candidates.append(m)

            if len(value) >= 40 and ("sap" in value.lower() or "marca" in value.lower()):
                description_candidates.append(value)
            elif "\n" in value and len(value) >= 40:
                description_candidates.append(value)

            if "$" in value or re.search(r"\b\d{1,3}(?:\.\d{3}){1,4}\b", value):
                money = parse_money(value)
                if money and money > 1000:
                    price_candidates.append(money)

        if "descripcion" in row_text.lower() and len(row_vals) >= 2:
            long_cell = sorted(row_vals, key=lambda x: len(x[1]), reverse=True)[0][1]
            if len(long_cell) >= 30:
                description_candidates.append(long_cell)

    wb.close()

    file_model = model_from_filename(xlsx_path.name)
    model = file_model
    if not norm_model(model) and model_candidates:
        ranked = sorted(model_candidates, key=lambda x: (len(x), x.count("/"), x), reverse=True)
        if norm_model(ranked[0]):
            model = ranked[0]

    description = ""
    if description_candidates:
        description = max(description_candidates, key=lambda x: len(x)).strip()
        description = re.sub(r"\n{3,}", "\n\n", description)

    price_cop = max(price_candidates) if price_candidates else None

    return {
        "model": model.strip(),
        "model_key": norm_model(model),
        "file_model": file_model,
        "description": description,
        "price_cop": price_cop,
        "xlsx_file": xlsx_path.name,
    }


def request_ok(resp: requests.Response):
    if 200 <= resp.status_code < 300:
        return
    raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:700]}")


def get_catalog_rows(base_url: str, service_key: str, owner_id: str, provider: str, timeout_sec: float):
    url = f"{base_url}/rest/v1/agent_product_catalog"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
    }
    params = {
        "select": "id,name,summary,description,source_payload,base_price_usd,price_currency,datasheet_url",
        "created_by": f"eq.{owner_id}",
        "provider": f"eq.{provider}",
        "limit": "10000",
    }
    resp = requests.get(url, headers=headers, params=params, timeout=(10, timeout_sec))
    request_ok(resp)
    return resp.json() if resp.text else []


def patch_catalog_row(base_url: str, service_key: str, row_id: str, patch_obj: dict, timeout_sec: float):
    url = f"{base_url}/rest/v1/agent_product_catalog"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    params = {"id": f"eq.{row_id}"}
    resp = requests.patch(
        url,
        headers=headers,
        params=params,
        data=json.dumps(patch_obj, ensure_ascii=False).encode("utf-8"),
        timeout=(10, timeout_sec),
    )
    request_ok(resp)


def build_folder_index(folder: Path):
    files = sorted(folder.glob("*.xlsx"))
    by_key = {}
    parsed = 0
    for file_path in files:
        data = parse_xlsx_product(file_path)
        key = data.get("model_key")
        if not key:
            continue
        parsed += 1
        prev = by_key.get(key)
        if not prev:
            by_key[key] = data
            continue
        prev_score = len(str(prev.get("description") or "")) + (1000 if prev.get("price_cop") else 0)
        next_score = len(str(data.get("description") or "")) + (1000 if data.get("price_cop") else 0)
        if next_score > prev_score:
            by_key[key] = data
    return by_key, len(files), parsed


def main():
    load_env()

    parser = argparse.ArgumentParser(description="Sync Ohaus Cotizaciones XLSX folder into agent_product_catalog")
    parser.add_argument("--folder", default=os.getenv("OHAUS_COTIZACIONES_FOLDER", DEFAULT_FOLDER))
    parser.add_argument("--owner-id", default=os.getenv("QUOTE_TEMPLATE_OWNER_ID", SYSTEM_USER_ID))
    parser.add_argument("--provider", default=os.getenv("QUOTE_TEMPLATE_PROVIDER", DEFAULT_PROVIDER))
    parser.add_argument("--trm", type=float, default=float(os.getenv("OHAUS_SYNC_TRM") or DEFAULT_TRM))
    parser.add_argument("--http-timeout", type=float, default=float(os.getenv("OHAUS_SYNC_HTTP_TIMEOUT") or 35))
    parser.add_argument("--skip-db", action="store_true", help="Only parse XLSX folder and show summary")
    parser.add_argument("--apply", action="store_true", help="Apply updates to DB. Default is dry-run")
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.is_absolute():
        folder = Path.cwd() / folder
    if not folder.exists():
        raise RuntimeError(f"Folder not found: {folder}")

    by_key, total_xlsx, parsed_xlsx = build_folder_index(folder)
    if not by_key:
        raise RuntimeError("No usable XLSX rows parsed from folder")

    print(f"Folder: {folder}")
    print(f"XLSX found: {total_xlsx}")
    print(f"XLSX parsed: {parsed_xlsx}")
    print(f"Unique models parsed: {len(by_key)}")

    if args.skip_db:
        print("Skip DB enabled. Parsing completed.")
        return

    base_url = (os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL") or "").strip().rstrip("/")
    service_key = (os.getenv("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
    if not base_url or not service_key:
        raise RuntimeError("Missing SUPABASE_URL/NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")

    print("Connecting to Supabase catalog...")
    catalog = get_catalog_rows(base_url, service_key, args.owner_id, args.provider, args.http_timeout)
    updates = []
    misses = 0

    now_iso = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

    for row in catalog:
        model = str(row.get("name") or "").strip()
        key = norm_model(model)
        incoming = by_key.get(key)
        if not incoming:
            misses += 1
            continue

        description = norm_text(incoming.get("description"))
        price_cop = incoming.get("price_cop")

        source_payload = row.get("source_payload") if isinstance(row.get("source_payload"), dict) else {}
        prices_cop = source_payload.get("prices_cop") if isinstance(source_payload.get("prices_cop"), dict) else {}
        merged_prices = dict(prices_cop)
        if price_cop:
            merged_prices["bogota"] = price_cop

        merged_source = {
            **source_payload,
            "import_source": "Ohaus/Cotizaciones",
            "quote_xlsx_file": incoming.get("xlsx_file"),
            "quote_model": incoming.get("model") or source_payload.get("quote_model") or model,
            "quote_description": description or source_payload.get("quote_description") or "",
            "descripcion_comercial_larga": description or source_payload.get("descripcion_comercial_larga") or "",
            "prices_cop": merged_prices,
            "cotizaciones_synced_at": now_iso,
        }

        patch_obj = {
            "source_payload": merged_source,
        }

        if description:
            patch_obj["description"] = description
            patch_obj["summary"] = description.replace("\n", " ")[:500]

        if price_cop and args.trm > 0:
            patch_obj["base_price_usd"] = round(price_cop / args.trm, 6)
            patch_obj["price_currency"] = "USD"
            patch_obj["last_price_update"] = now_iso

        updates.append((str(row.get("id")), model, patch_obj, incoming))

    with_price = sum(1 for _, _, _, inc in updates if inc.get("price_cop"))
    with_desc = sum(1 for _, _, patch, _ in updates if patch.get("description"))

    print(f"Catalog rows: {len(catalog)}")
    print(f"Catalog rows matched: {len(updates)}")
    print(f"Matched with description: {with_desc}")
    print(f"Matched with price: {with_price}")
    print(f"Catalog rows without match: {misses}")

    if updates:
        sample = updates[0]
        sample_inc = sample[3]
        print(
            "Sample:",
            json.dumps(
                {
                    "catalog_model": sample[1],
                    "xlsx_file": sample_inc.get("xlsx_file"),
                    "price_cop": sample_inc.get("price_cop"),
                    "desc_len": len(str(sample_inc.get("description") or "")),
                },
                ensure_ascii=False,
            ),
        )

    if not args.apply:
        print("Dry run mode. Use --apply to write changes.")
        return

    for row_id, _model, patch_obj, _incoming in updates:
        patch_catalog_row(base_url, service_key, row_id, patch_obj, args.http_timeout)

    print(f"Updated rows: {len(updates)}")


if __name__ == "__main__":
    main()
