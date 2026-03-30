import json
import os
import re
from pathlib import Path

import openpyxl
import requests


SYSTEM_TENANT_ID = "0811c118-5a2f-40cb-907e-8979e0984096"
SYSTEM_USER_ID = "841263c6-196d-49cd-b5ba-aae0b097014f"
DEFAULT_CLIENT_XLSX = "app/api/agents/channels/evolution/webhook-v2/Contactos Nuevo CRM Cliente.xlsx"
DEFAULT_DISTRIBUTOR_XLSX = "app/api/agents/channels/evolution/webhook-v2/Contactos Nuevo CRM distribuidor.xlsx"


def load_env_file(path: Path):
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and k not in os.environ:
            os.environ[k] = v


def load_env():
    root = Path.cwd()
    load_env_file(root / ".env.local")
    load_env_file(root / ".env")


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def normalize_key(text: str) -> str:
    t = normalize_text(text).lower()
    return re.sub(r"[^a-z0-9]+", "_", t).strip("_")


def clean_value(v):
    s = normalize_text(str(v or ""))
    if s in {"-", "--", "n/a", "na", "none", "null"}:
        return ""
    return s


def normalize_phone(v: str) -> str:
    digits = re.sub(r"\D+", "", v or "")
    if not digits:
        return ""
    if digits.startswith("57") and len(digits) >= 12:
        return digits
    if len(digits) == 10:
        return f"57{digits}"
    return digits


def city_tier_from_department(dept: str) -> str:
    t = normalize_key(dept)
    if "bogota" in t:
        return "bogota"
    if "antioquia" in t:
        return "antioquia"
    return "bogota"


def parse_rows(xlsx_path: Path, customer_type: str, tenant_id: str, created_by: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [clean_value(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {headers[idx]: clean_value(row[idx]) for idx in range(min(len(headers), len(row)))}

        name = clean_value(data.get("Contactos Nombre y apellido", ""))
        phone = normalize_phone(clean_value(data.get("Contactos Celular ", "")))
        email = clean_value(data.get("Contactos Correo", "")).lower()
        company = clean_value(data.get("Contactos Empresa (si aplica)", ""))
        nit = re.sub(r"[^0-9\-]", "", clean_value(data.get("Empresas NIT", "")))
        department = clean_value(data.get("Departamento", ""))
        address = clean_value(data.get("Empresas Dirección", "")) or clean_value(data.get("Empresas Direccion", "")) or clean_value(data.get("Empresas Direcci�n", ""))
        activity = clean_value(data.get("Actividad del cliente", ""))
        assigned_to = clean_value(data.get("Contactos Asignado a", ""))

        contact_key = ""
        if phone:
            contact_key = phone
        elif nit:
            contact_key = f"nit:{nit}"
        elif email:
            contact_key = f"email:{email}"
        elif name and company:
            contact_key = f"name:{normalize_key(name)}:{normalize_key(company)}"
        else:
            continue

        is_distributor = customer_type == "distributor"
        price_tier = "distribuidor" if is_distributor else city_tier_from_department(department)
        metadata = {
            "customer_type": customer_type,
            "price_tier": price_tier,
            "nit": nit,
            "billing_city": department,
            "address": address,
            "activity": activity,
            "assigned_to": assigned_to,
            "source": "xlsx_crm_contacts_import",
            "source_file": xlsx_path.name,
            "row_number": i,
        }

        rows.append(
            {
                "tenant_id": tenant_id,
                "created_by": created_by,
                "contact_key": contact_key,
                "name": name or None,
                "email": email or None,
                "phone": phone or None,
                "company": company or None,
                "status": "analysis",
                "metadata": metadata,
            }
        )

    return rows


def request_ok(resp: requests.Response):
    if 200 <= resp.status_code < 300:
        return
    raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:600]}")


def is_status_constraint_error(resp: requests.Response) -> bool:
    if resp is None:
        return False
    msg = (resp.text or "").lower()
    return resp.status_code == 400 and "status_check" in msg and "agent_crm_contacts" in msg


def chunked(arr, size):
    for i in range(0, len(arr), size):
        yield arr[i : i + size]


def upsert_contacts(base_url: str, service_key: str, rows):
    url = f"{base_url}/rest/v1/agent_crm_contacts"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal,resolution=merge-duplicates",
    }
    params = {"on_conflict": "created_by,contact_key"}

    total = 0
    for batch in chunked(rows, 300):
        resp = requests.post(url, headers=headers, params=params, data=json.dumps(batch, ensure_ascii=False).encode("utf-8"), timeout=120)
        if is_status_constraint_error(resp):
            legacy_batch = [{**r, "status": "draft"} for r in batch]
            resp = requests.post(url, headers=headers, params=params, data=json.dumps(legacy_batch, ensure_ascii=False).encode("utf-8"), timeout=120)
        request_ok(resp)
        total += len(batch)
    return total


def dedupe_contacts(rows):
    merged = {}
    for row in rows:
        key = f"{row.get('created_by','')}::{row.get('contact_key','')}"
        if key not in merged:
            merged[key] = row
            continue
        cur = merged[key]
        for fld in ["name", "email", "phone", "company"]:
            if not cur.get(fld) and row.get(fld):
                cur[fld] = row.get(fld)

        cur_meta = dict(cur.get("metadata") or {})
        new_meta = dict(row.get("metadata") or {})
        if new_meta.get("customer_type") == "distributor":
            cur_meta["customer_type"] = "distributor"
            cur_meta["price_tier"] = "distribuidor"
        for k, v in new_meta.items():
            if v and not cur_meta.get(k):
                cur_meta[k] = v
        cur["metadata"] = cur_meta
        merged[key] = cur
    return list(merged.values())


def main():
    load_env()

    supabase_url = (os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL") or "").strip().rstrip("/")
    service_key = (os.getenv("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
    tenant_id = (os.getenv("CRM_IMPORT_TENANT_ID") or SYSTEM_TENANT_ID).strip()
    created_by = (os.getenv("CRM_IMPORT_USER_ID") or SYSTEM_USER_ID).strip()
    dry_run = (os.getenv("CRM_IMPORT_DRY_RUN") or "false").strip().lower() in {"1", "true", "yes"}

    client_xlsx = Path(os.getenv("CRM_IMPORT_CLIENT_XLSX") or DEFAULT_CLIENT_XLSX)
    dist_xlsx = Path(os.getenv("CRM_IMPORT_DISTRIBUTOR_XLSX") or DEFAULT_DISTRIBUTOR_XLSX)
    if not client_xlsx.is_absolute():
        client_xlsx = Path.cwd() / client_xlsx
    if not dist_xlsx.is_absolute():
        dist_xlsx = Path.cwd() / dist_xlsx

    if not client_xlsx.exists():
        raise RuntimeError(f"Client XLSX not found: {client_xlsx}")
    if not dist_xlsx.exists():
        raise RuntimeError(f"Distributor XLSX not found: {dist_xlsx}")

    client_rows = parse_rows(client_xlsx, "client", tenant_id, created_by)
    dist_rows = parse_rows(dist_xlsx, "distributor", tenant_id, created_by)
    rows = dedupe_contacts(client_rows + dist_rows)

    print(f"Client rows parsed: {len(client_rows)}")
    print(f"Distributor rows parsed: {len(dist_rows)}")
    print(f"Total rows parsed: {len(rows)}")
    if rows:
        print("Sample:", json.dumps(rows[:2], ensure_ascii=False)[:900])

    if dry_run:
        print("Dry run enabled, no DB write executed.")
        return

    if not supabase_url or not service_key:
        raise RuntimeError("Missing SUPABASE_URL/NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")

    upserted = upsert_contacts(supabase_url, service_key, rows)
    print(f"Upserted rows: {upserted}")


if __name__ == "__main__":
    main()
